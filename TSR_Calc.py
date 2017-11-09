#This Python code is to extract data from Excel Sheet using OpenPyXl package in Python 3.6
import openpyxl as xl

wb=xl.load_workbook("E:\\Harshitha_work\\CEO Remuneration Work\\Additional Work\\Returns Calc\Datastream Collection Entire Dataset 170911.xlsx")

all_sheets=wb.get_sheet_names()

a=[]
b=[]

for sheet in all_sheets:
    #print("sheet name is: ", sheet)
    sh=wb.get_sheet_by_name(sheet)
    if sh['A3'].value is not None:
        #print("Total Columns in the sheet are: ",sh.max_column)
        for i in range(sh.max_column):
            b=[]
            b.append(sheet)
            b.append(sh.cell(row=3,column=i+1).value)
            b.append(sh.cell(row=4,column=i+1).value)
            b.append(3)
            b.append(i+1)
            a.append(b)
    else:
          #print("Total Columns in the sheet are: ",sh.max_column)
          for i in range(sh.max_column):
            b=[]
            b.append(sheet)
            b.append(sh.cell(row=4,column=i+1).value)
            b.append(sh.cell(row=5,column=i+1).value)
            b.append(4)
            b.append(i+1)
            a.append(b)

calc='y'

while calc.lower()=='y':
                print("Please Enter Company Name first word")
                company_name=input()
                print("Please enter month of join of CEO")
                mm_st=int(input())
                print("Please enter year of join of CEO")
                yy_st=int(input())


                for c in a:
                                if c[1] is not None:
                                                if company_name in c[1].lower():
                                                                print(c)
                                                                cont_y_n=input()
                                                                if cont_y_n.lower() == 'y':
                                                                                a_sh=wb.get_sheet_by_name(c[0])
                                                                                for x in range(601):
                                                                                                dd=a_sh.cell(row=c[3]+2+x,column=1).value
                                                                                                #print(dd)
                                                                                                if ((mm_st == dd.month) and (yy_st == dd.year)):
                                                                                                                st_price=str(a_sh.cell(row=c[3]+2+x,column=c[4]).value)
                                                                                                                en_price=str(a_sh.cell(row=c[3]+2+x+36,column=c[4]).value)
                                                                                                                if (st_price == 'NA') or (en_price == 'NA'):
                                                                                                                                print("Start Price or End Price is NA")
                                                                                                                                print("Start Price is: ",st_price," End Price is: ",en_price) 
                                                                                                                else:
                                                                                                                                print("TS Start Value is: ", a_sh.cell(row=c[3]+2+x,column=c[4]).value)
                                                                                                                                print("TS End Value is : ",a_sh.cell(row=c[3]+2+x+36,column=c[4]).value)
                                                                                                                                TSR=(a_sh.cell(row=c[3]+2+x+36,column=c[4]).value - a_sh.cell(row=c[3]+2+x,column=c[4]).value)/(a_sh.cell(row=c[3]+2+x,column=c[4]).value)
                                                                                                                                print("The TSR percentage is : ", TSR*100)
                                                                                                

                print("Do you want to continue calculating TSR for others:")
                calc=input()
                
